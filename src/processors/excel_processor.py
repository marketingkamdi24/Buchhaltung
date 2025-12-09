"""
Excel processor module for eBay transaction reports.
Processes Excel files according to specified rules, matching the exact structure
of the Gradio-based processor with KD-NR and RG-NR mapping from API data.
"""
import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime
from dataclasses import dataclass, field
from typing import Optional, List, Dict, Any
import os
import sys
import pandas as pd

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
from config import get_config, get_output_path


@dataclass
class ProcessingResult:
    success: bool
    message: str
    file_path: Optional[str] = None
    task_results: List[str] = field(default_factory=list)


def find_header_row(ws):
    """Find the row containing column headers by looking for 'Datum der Transaktionserstellung'"""
    for row_idx in range(1, min(20, ws.max_row + 1)):
        for col_idx in range(1, min(10, ws.max_column + 1)):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val and 'Datum der Transaktionserstellung' in str(val):
                return row_idx
    return None


def get_column_mapping(ws, header_row):
    """Create mapping from column name to column index"""
    mapping = {}
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col_idx).value
        if val:
            mapping[str(val).strip()] = col_idx
    return mapping


def get_numeric_value(val):
    """Get numeric value from cell, converting '--' to 0"""
    if val == '--' or val == '—' or val is None:
        return 0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0


def format_date_yyyymm(date_val):
    """Format date as YYYYMM string"""
    if isinstance(date_val, datetime):
        return date_val.strftime("%Y%m")
    elif isinstance(date_val, str):
        try:
            dt = datetime.strptime(date_val[:10], "%Y-%m-%d")
            return dt.strftime("%Y%m")
        except:
            pass
    return ""


class ExcelProcessor:
    """Process Excel files according to eBay transaction report rules."""
    
    def __init__(self):
        self.config = get_config().excel
        self.results = []
    
    def _log(self, message: str):
        """Add a message to the processing log."""
        self.results.append(message)
    
    def process(self, file_path: str, api_data: Optional[pd.DataFrame] = None) -> ProcessingResult:
        """
        Process the Excel file according to specified rules.
        
        Args:
            file_path: Path to the Excel file to process
            api_data: Optional DataFrame with API data for KD-NR/RG-NR mapping
                     Must contain ORDER_ID, KUNDENNR, and BELEGNR columns
        
        Returns:
            ProcessingResult with success status, message, and output file path
        """
        self.results = []
        output_file = None
        
        try:
            # Load the workbook
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            original_sheet_name = ws.title
            
            self._log(f"📄 Loaded file: {os.path.basename(file_path)}")
            self._log(f"   Original dimensions: {ws.max_row} rows × {ws.max_column} columns")
            
            # Build ORDER_ID to KUNDENNR/BELEGNR mapping from API data
            kdnr_mapping = {}  # Maps Bestellnummer -> KUNDENNR
            rgnr_mapping = {}  # Maps Bestellnummer -> BELEGNR
            
            if api_data is not None and not api_data.empty:
                if 'ORDER_ID' in api_data.columns:
                    if 'KUNDENNR' in api_data.columns:
                        for _, row in api_data.iterrows():
                            order_id = str(row['ORDER_ID']) if pd.notna(row['ORDER_ID']) else None
                            kundennr = row['KUNDENNR'] if pd.notna(row.get('KUNDENNR')) else None
                            if order_id and kundennr:
                                kdnr_mapping[order_id] = kundennr
                        self._log(f"✅ Loaded {len(kdnr_mapping)} KUNDENNR mappings from API data")
                    
                    if 'BELEGNR' in api_data.columns:
                        for _, row in api_data.iterrows():
                            order_id = str(row['ORDER_ID']) if pd.notna(row['ORDER_ID']) else None
                            belegnr = row['BELEGNR'] if pd.notna(row.get('BELEGNR')) else None
                            if order_id and belegnr:
                                rgnr_mapping[order_id] = belegnr
                        self._log(f"✅ Loaded {len(rgnr_mapping)} BELEGNR mappings from API data")
            
            # Step 1: Find the header row
            header_row = find_header_row(ws)
            if not header_row:
                self._log("❌ Could not find header row with 'Datum der Transaktionserstellung'")
                return ProcessingResult(
                    success=False,
                    message="\n".join(self.results),
                    task_results=self.results.copy()
                )
            self._log(f"✅ Found header row at row {header_row}")
            
            # Step 2: Get column mapping from original file
            col_mapping = get_column_mapping(ws, header_row)
            self._log(f"✅ Found {len(col_mapping)} columns in header")
            
            # Step 3: Extract metadata from rows before header
            verkaeufer_value = None
            betrag_value = None
            
            for row_idx in range(1, header_row):
                a_val = ws.cell(row=row_idx, column=1).value
                b_val = ws.cell(row=row_idx, column=2).value
                if a_val and 'Verkäufer' in str(a_val):
                    verkaeufer_value = b_val
                if a_val and 'Betrag' in str(a_val):
                    betrag_value = b_val
            
            self._log(f"✅ Extracted metadata: Verkäufer={verkaeufer_value}, Betrag={betrag_value}")
            
            # Step 4: Read all data rows and separate "Andere Gebühr" from regular transactions
            data_rows = []
            andere_gebuehr_rows = []
            
            # Get necessary column indices
            typ_col = col_mapping.get('Typ', 2)
            datum_col = col_mapping.get('Datum der Transaktionserstellung', 1)
            bestellnr_col = col_mapping.get('Bestellnummer', 3)
            alte_bestellnr_col = col_mapping.get('Alte Bestellnummer', 4)
            nutzername_col = col_mapping.get('Nutzername des Käufers', 5)
            name_col = col_mapping.get('Name des Käufers', 6)
            betrag_abzgl_col = col_mapping.get('Betrag abzügl. Kosten')
            auszahlungsdatum_col = col_mapping.get('Auszahlungsdatum')
            auszahlungsnr_col = col_mapping.get('Auszahlung Nr.')
            transaktionsbetrag_col = col_mapping.get('Transaktionsbetrag (inkl. Kosten)')
            zwischensumme_col = col_mapping.get('Zwischensumme Artikel')
            verpackung_col = col_mapping.get('Verpackung und Versand')
            fixer_anteil_col = col_mapping.get('Fixer Anteil der Verkaufsprovision')
            variabler_anteil_col = col_mapping.get('Variabler Anteil der Verkaufsprovision')
            gebuehr_hohe_quote_col = col_mapping.get('Gebühr für sehr hohe Quote an „nicht wie beschriebenen Artikeln"')
            gebuehr_servicestatus_col = col_mapping.get('Gebühr für unterdurchschnittlichen Servicestatus')
            internationale_gebuehr_col = col_mapping.get('Internationale Gebühr')
            
            for row_idx in range(header_row + 1, ws.max_row + 1):
                typ = ws.cell(row=row_idx, column=typ_col).value
                if not typ:
                    continue
                
                # Get Zwischensumme and add Verpackung und Versand if exists
                zwischensumme = get_numeric_value(ws.cell(row=row_idx, column=zwischensumme_col).value) if zwischensumme_col else 0
                verpackung = get_numeric_value(ws.cell(row=row_idx, column=verpackung_col).value) if verpackung_col else 0
                if verpackung != 0:
                    zwischensumme += verpackung
                
                bestellnummer = ws.cell(row=row_idx, column=bestellnr_col).value
                
                row_data = {
                    'datum': ws.cell(row=row_idx, column=datum_col).value,
                    'typ': typ,
                    'bestellnummer': bestellnummer,
                    'alte_bestellnummer': ws.cell(row=row_idx, column=alte_bestellnr_col).value if alte_bestellnr_col else None,
                    'nutzername': ws.cell(row=row_idx, column=nutzername_col).value,
                    'name': ws.cell(row=row_idx, column=name_col).value,
                    'transaktionsbetrag': get_numeric_value(ws.cell(row=row_idx, column=transaktionsbetrag_col).value) if transaktionsbetrag_col else 0,
                    'zwischensumme': zwischensumme,
                    'fixer_anteil': get_numeric_value(ws.cell(row=row_idx, column=fixer_anteil_col).value) if fixer_anteil_col else 0,
                    'variabler_anteil': get_numeric_value(ws.cell(row=row_idx, column=variabler_anteil_col).value) if variabler_anteil_col else 0,
                    'gebuehr_hohe_quote': get_numeric_value(ws.cell(row=row_idx, column=gebuehr_hohe_quote_col).value) if gebuehr_hohe_quote_col else 0,
                    'gebuehr_servicestatus': get_numeric_value(ws.cell(row=row_idx, column=gebuehr_servicestatus_col).value) if gebuehr_servicestatus_col else 0,
                    'internationale_gebuehr': get_numeric_value(ws.cell(row=row_idx, column=internationale_gebuehr_col).value) if internationale_gebuehr_col else 0,
                    'betrag_abzgl': get_numeric_value(ws.cell(row=row_idx, column=betrag_abzgl_col).value) if betrag_abzgl_col else 0,
                    'auszahlungsnr': ws.cell(row=row_idx, column=auszahlungsnr_col).value if auszahlungsnr_col else None,
                    'auszahlungsdatum': ws.cell(row=row_idx, column=auszahlungsdatum_col).value if auszahlungsdatum_col else None,
                    # Map KD-NR and RG-NR from API data using Bestellnummer
                    'kdnr': kdnr_mapping.get(str(bestellnummer)) if bestellnummer else None,
                    'rgnr': rgnr_mapping.get(str(bestellnummer)) if bestellnummer else None,
                }
                
                if 'Andere Geb' in str(typ):
                    andere_gebuehr_rows.append(row_data)
                else:
                    data_rows.append(row_data)
            
            self._log(f"✅ Found {len(data_rows)} regular transactions")
            self._log(f"✅ Found {len(andere_gebuehr_rows)} 'Andere Gebühr' transactions")
            
            # Count matched KD-NR and RG-NR
            kdnr_matched = sum(1 for r in data_rows if r['kdnr'] is not None)
            rgnr_matched = sum(1 for r in data_rows if r['rgnr'] is not None)
            if kdnr_mapping or rgnr_mapping:
                self._log(f"✅ Matched KD-NR for {kdnr_matched}/{len(data_rows)} transactions")
                self._log(f"✅ Matched RG-NR for {rgnr_matched}/{len(data_rows)} transactions")
            
            # Step 5: Create new workbook with processed data
            new_wb = openpyxl.Workbook()
            new_ws = new_wb.active
            new_ws.title = original_sheet_name
            
            # Row 1: Verkäufer
            new_ws['A1'] = 'Verkäufer'
            new_ws['B1'] = verkaeufer_value
            
            # Row 2: Transaktionen count
            new_ws['A2'] = 'Transaktionen'
            new_ws['B2'] = len(data_rows) + len(andere_gebuehr_rows)
            
            # Row 3: Betrag
            new_ws['A3'] = 'Betrag'
            new_ws['B3'] = betrag_value
            
            # Row 4: Headers (A-R main columns, T and U for verification)
            output_headers = [
                'Datum der Transaktionserstellung',  # A
                'Typ',  # B
                'Bestellnummer',  # C
                'Alte Bestellnummer',  # D
                'Nutzername des Käufers',  # E
                'Name des Käufers',  # F
                'KD-NR',  # G
                'RG-NR',  # H
                'Transaktionsbetrag (inkl. Kosten)',  # I
                'Zwischensumme Artikel',  # J
                'Fixer Anteil der Verkaufsprovision',  # K
                'Variabler Anteil der Verkaufsprovision',  # L
                'Gebühr für sehr hohe Quote an „nicht wie beschriebenen Artikeln"',  # M
                'Gebühr für unterdurchschnittlichen Servicestatus',  # N
                'Internationale Gebühr',  # O
                'Betrag abzügl. Kosten',  # P
                'Auszahlung Nr.',  # Q
                'Auszahlungsdatum',  # R
            ]
            
            for col_idx, header in enumerate(output_headers, 1):
                new_ws.cell(row=4, column=col_idx, value=header)
            
            # Row 5+: Data rows
            current_row = 5
            last_date = None
            auszahlungsnr = None
            auszahlungsdatum = None
            
            for row_data in data_rows:
                new_ws.cell(row=current_row, column=1, value=row_data['datum'])  # A
                new_ws.cell(row=current_row, column=2, value=row_data['typ'])  # B
                new_ws.cell(row=current_row, column=3, value=row_data['bestellnummer'])  # C
                new_ws.cell(row=current_row, column=4, value=row_data['alte_bestellnummer'])  # D
                new_ws.cell(row=current_row, column=5, value=row_data['nutzername'])  # E
                new_ws.cell(row=current_row, column=6, value=row_data['name'])  # F
                new_ws.cell(row=current_row, column=7, value=row_data['kdnr'])  # G - KD-NR from API
                new_ws.cell(row=current_row, column=8, value=row_data['rgnr'])  # H - RG-NR from API
                new_ws.cell(row=current_row, column=9, value=row_data['transaktionsbetrag'])  # I
                new_ws.cell(row=current_row, column=10, value=row_data['zwischensumme'])  # J
                new_ws.cell(row=current_row, column=11, value=row_data['fixer_anteil'])  # K
                new_ws.cell(row=current_row, column=12, value=row_data['variabler_anteil'])  # L
                new_ws.cell(row=current_row, column=13, value=row_data['gebuehr_hohe_quote'])  # M
                new_ws.cell(row=current_row, column=14, value=row_data['gebuehr_servicestatus'])  # N
                new_ws.cell(row=current_row, column=15, value=row_data['internationale_gebuehr'])  # O
                new_ws.cell(row=current_row, column=16, value=row_data['betrag_abzgl'])  # P
                new_ws.cell(row=current_row, column=17, value=row_data['auszahlungsnr'])  # Q
                new_ws.cell(row=current_row, column=18, value=row_data['auszahlungsdatum'])  # R
                # S - empty/small column
                new_ws.cell(row=current_row, column=20, value=row_data['zwischensumme'])  # T - Copy of Zwischensumme for verification
                new_ws.cell(row=current_row, column=21, value=f"=J{current_row}-T{current_row}")  # U - Verification formula
                
                # Track values for later use
                if row_data['datum']:
                    last_date = row_data['datum']
                if row_data['auszahlungsnr']:
                    auszahlungsnr = row_data['auszahlungsnr']
                if row_data['auszahlungsdatum']:
                    auszahlungsdatum = row_data['auszahlungsdatum']
                
                current_row += 1
            
            data_end_row = current_row - 1
            self._log(f"✅ Wrote {len(data_rows)} data rows (rows 5-{data_end_row})")
            
            # Get date for "Gebuehr" row
            gebuehr_date = last_date
            if andere_gebuehr_rows:
                andere_gebuehr_date = andere_gebuehr_rows[0]['datum']
            else:
                andere_gebuehr_date = last_date
            
            # Add "Gebuehr" row with formulas
            gebuehr_row = current_row
            new_ws.cell(row=gebuehr_row, column=1, value=gebuehr_date)  # A
            new_ws.cell(row=gebuehr_row, column=6, value="Gebuehr")  # F
            new_ws.cell(row=gebuehr_row, column=7, value="7400700")  # G
            new_ws.cell(row=gebuehr_row, column=8, value=format_date_yyyymm(gebuehr_date))  # H
            new_ws.cell(row=gebuehr_row, column=9, value=f"=J{gebuehr_row}")  # I
            new_ws.cell(row=gebuehr_row, column=10, value=f"=SUM(K{gebuehr_row}:O{gebuehr_row})")  # J
            new_ws.cell(row=gebuehr_row, column=11, value=f"=SUM(K5:K{data_end_row})")  # K
            new_ws.cell(row=gebuehr_row, column=12, value=f"=SUM(L5:L{data_end_row})")  # L
            new_ws.cell(row=gebuehr_row, column=13, value=f"=SUM(M5:M{data_end_row})")  # M
            new_ws.cell(row=gebuehr_row, column=14, value=f"=SUM(N5:N{data_end_row})")  # N
            new_ws.cell(row=gebuehr_row, column=15, value=f"=SUM(O5:O{data_end_row})")  # O
            new_ws.cell(row=gebuehr_row, column=16, value=0)  # P
            new_ws.cell(row=gebuehr_row, column=17, value=auszahlungsnr)  # Q
            new_ws.cell(row=gebuehr_row, column=18, value=auszahlungsdatum)  # R
            self._log(f"✅ Added 'Gebuehr' row at row {gebuehr_row}")
            
            # Add "Andere Gebuehr" summary row
            andere_gebuehr_row = current_row + 1
            andere_gebuehr_sum = sum(row['betrag_abzgl'] for row in andere_gebuehr_rows)
            
            new_ws.cell(row=andere_gebuehr_row, column=1, value=andere_gebuehr_date)  # A
            new_ws.cell(row=andere_gebuehr_row, column=6, value="Andere Gebuehr")  # F
            new_ws.cell(row=andere_gebuehr_row, column=7, value="7400700")  # G
            new_ws.cell(row=andere_gebuehr_row, column=8, value=format_date_yyyymm(andere_gebuehr_date))  # H
            new_ws.cell(row=andere_gebuehr_row, column=9, value=f"=J{andere_gebuehr_row}")  # I
            new_ws.cell(row=andere_gebuehr_row, column=10, value=andere_gebuehr_sum)  # J
            new_ws.cell(row=andere_gebuehr_row, column=16, value=andere_gebuehr_sum)  # P
            new_ws.cell(row=andere_gebuehr_row, column=17, value=auszahlungsnr)  # Q
            new_ws.cell(row=andere_gebuehr_row, column=18, value=auszahlungsdatum)  # R
            self._log(f"✅ Added 'Andere Gebuehr' row at row {andere_gebuehr_row} (sum: {andere_gebuehr_sum})")
            
            # Add totals row
            totals_row = current_row + 2
            new_ws.cell(row=totals_row, column=10, value=f"=SUM(J5:J{andere_gebuehr_row})")  # J
            new_ws.cell(row=totals_row, column=16, value=f"=SUM(P5:P{andere_gebuehr_row})")  # P
            self._log(f"✅ Added totals row at row {totals_row}")
            
            # Step 6: Create second sheet with Andere Gebühr details
            if andere_gebuehr_rows:
                andere_ws = new_wb.create_sheet("Tabelle1")
                
                for idx, row_data in enumerate(andere_gebuehr_rows, 1):
                    andere_ws.cell(row=idx, column=1, value=row_data['datum'])  # A
                    andere_ws.cell(row=idx, column=2, value=row_data['typ'])  # B
                    andere_ws.cell(row=idx, column=3, value=row_data['bestellnummer'])  # C
                    andere_ws.cell(row=idx, column=4, value=row_data['alte_bestellnummer'])  # D
                    andere_ws.cell(row=idx, column=5, value=row_data['nutzername'])  # E
                    andere_ws.cell(row=idx, column=6, value=row_data['name'])  # F
                    andere_ws.cell(row=idx, column=7, value="7400700")  # G
                    andere_ws.cell(row=idx, column=8, value=format_date_yyyymm(row_data['datum']))  # H
                    andere_ws.cell(row=idx, column=9, value=row_data['betrag_abzgl'])  # I
                    andere_ws.cell(row=idx, column=10, value=row_data['betrag_abzgl'])  # J
                    andere_ws.cell(row=idx, column=11, value=0)  # K
                    andere_ws.cell(row=idx, column=12, value=0)  # L
                    andere_ws.cell(row=idx, column=13, value=0)  # M
                    andere_ws.cell(row=idx, column=14, value=0)  # N
                    andere_ws.cell(row=idx, column=15, value=0)  # O
                    andere_ws.cell(row=idx, column=16, value=row_data['betrag_abzgl'])  # P
                    andere_ws.cell(row=idx, column=17, value=row_data['auszahlungsnr'])  # Q
                    andere_ws.cell(row=idx, column=18, value=row_data['auszahlungsdatum'])  # R
                
                # Add empty row and SUM formula for Tabelle1
                last_data_row = len(andere_gebuehr_rows)
                sum_row = last_data_row + 2  # Empty row + 1
                # SUM formula should include the empty row (J1:J{sum_row-1})
                andere_ws.cell(row=sum_row, column=10, value=f"=SUM(J1:J{sum_row-1})")  # J SUM
                andere_ws.cell(row=sum_row, column=16, value=f"=SUM(P1:P{sum_row-1})")  # P SUM
                
                self._log(f"✅ Created 'Tabelle1' sheet with {len(andere_gebuehr_rows)} 'Andere Gebühr' entries + SUM formula")
            
            # Save the workbook
            original_name = os.path.basename(file_path)
            output_filename = f"bearbeitet_{original_name}"
            output_path = get_output_path(output_filename)
            new_wb.save(output_path)
            output_file = str(output_path)
            
            self._log(f"\n✅ File saved successfully as '{output_filename}'")
            self._log(f"\n📊 Summary:")
            self._log(f"   - Regular transactions: {len(data_rows)}")
            self._log(f"   - Andere Gebühr entries: {len(andere_gebuehr_rows)}")
            self._log(f"   - Total amount (Andere Gebühr): {andere_gebuehr_sum}")
            
            return ProcessingResult(
                success=True,
                message="\n".join(self.results),
                file_path=output_file,
                task_results=self.results.copy()
            )
            
        except Exception as e:
            self._log(f"\n❌ Critical Error: {str(e)}")
            import traceback
            self._log(f"\n{traceback.format_exc()}")
            return ProcessingResult(
                success=False,
                message="\n".join(self.results),
                task_results=self.results.copy()
            )


def process_excel(file, api_data: Optional[pd.DataFrame] = None):
    """
    Wrapper function to process Excel file.
    
    Args:
        file: File object or path to the Excel file
        api_data: Optional DataFrame with API data for KD-NR/RG-NR mapping
    
    Returns:
        Tuple of (message, file_path)
    """
    if file is None:
        return "Upload Excel file", None
    processor = ExcelProcessor()
    file_path = file.name if hasattr(file, 'name') else str(file)
    result = processor.process(file_path, api_data)
    return result.message, result.file_path
